import os
import re
import json
import uuid
import traceback
import logging
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from functools import wraps
from flask import Flask, request, render_template, send_file, jsonify, send_from_directory, session, redirect, url_for
from werkzeug.utils import secure_filename
from csv_parser import (
    parse_database_csv, parse_database_excel,
    parse_tax_csv, parse_tax_excel,
    Employee, TaxRecord
)
from pdf_generator import generate_pdf, fmt_bdt
from typing import Dict, List, Optional
from db_manager import (
    init_db, save_employees, save_tax_records,
    load_employees, load_tax_records, get_fiscal_years,
    get_all_dataset_batches, delete_batch, update_employee_tin,
    delete_fiscal_year, clear_all_data, get_connection,
    create_tax_request, get_tax_requests, get_tax_request_by_req_id, update_tax_request_status
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
log = logging.getLogger(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

DB_CSV_PATH = os.path.join(DATA_DIR, "database.csv")
DB_EXCEL_PATH = os.path.join(DATA_DIR, "database.xlsx")
TAX_CSV_PATH = os.path.join(DATA_DIR, "tax.csv")
TAX_EXCEL_PATH = os.path.join(DATA_DIR, "tax.xlsx")
META_PATH = os.path.join(DATA_DIR, "meta.json")

app = Flask(__name__)
app.secret_key = "income_tax_crm_super_secret_key_2026"
app.config["OUTPUT_FOLDER"] = OUTPUT_DIR
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024

db_store: Dict[str, Employee] = {}
tax_store: Dict[str, List[TaxRecord]] = {}

USERS_PATH = os.path.join(DATA_DIR, "users.json")

DEFAULT_USERS = {
    "admin": {
        "username": "admin",
        "password": "admin123",
        "name": "System Administrator",
        "role": "Super Admin",
        "department": "All Departments"
    },
    "accounts": {
        "username": "accounts",
        "password": "acc123",
        "name": "Finance & Payroll Officer",
        "role": "Finance Manager (Full Access)",
        "department": "All Departments"
    },
    "hr": {
        "username": "hr",
        "password": "hr123",
        "name": "Human Resources Officer",
        "role": "HR Manager (Full Access)",
        "department": "All Departments"
    },
    "brac_ied": {
        "username": "brac_ied",
        "password": "ied123",
        "name": "BRAC IED Officer",
        "role": "Department Manager",
        "department": "BRAC IED"
    },
    "brac_univ": {
        "username": "brac_univ",
        "password": "univ123",
        "name": "BRAC University Officer",
        "role": "Department Manager",
        "department": "BRAC University"
    }
}


def _load_users() -> dict:
    if os.path.exists(USERS_PATH):
        try:
            with open(USERS_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            log.error(f"Error reading users.json: {e}")
    _save_users(DEFAULT_USERS)
    return DEFAULT_USERS


def _save_users(users_dict: dict):
    try:
        with open(USERS_PATH, "w", encoding="utf-8") as f:
            json.dump(users_dict, f, indent=2)
    except Exception as e:
        log.error(f"Error saving users.json: {e}")


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user" not in session:
            if request.path.startswith("/api/"):
                return jsonify({"error": "Unauthorized. Please log in."}), 401
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        curr_user = session.get("user", {})
        if curr_user.get("role") != "Super Admin" and curr_user.get("username") != "admin":
            return jsonify({"error": "Admin privileges required"}), 403
        return f(*args, **kwargs)
    return decorated_function


def _load_meta():
    if os.path.exists(META_PATH):
        with open(META_PATH) as f:
            return json.load(f)
    return {}


def _save_meta(key, filename, count):
    meta = _load_meta()
    meta[key] = {
        "filename": filename,
        "count": count,
        "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    with open(META_PATH, "w") as f:
        json.dump(meta, f, indent=2)


def _clear_meta(key=None):
    if not key or key == "all":
        if os.path.exists(META_PATH):
            try:
                os.remove(META_PATH)
            except Exception:
                pass
    else:
        meta = _load_meta()
        if key in meta:
            del meta[key]
            with open(META_PATH, "w") as f:
                json.dump(meta, f, indent=2)


def _seed_db_if_empty():
    init_db()


def _get_stores(fiscal_year: Optional[str] = None):
    _seed_db_if_empty()
    if not fiscal_year:
        fiscal_year = request.args.get("fiscal_year")

    fys = get_fiscal_years()
    if not fiscal_year or fiscal_year.lower() in ('default', 'none', ''):
        fiscal_year = fys[0] if fys else "2025-2026"

    db = load_employees(fiscal_year if fiscal_year != "all" else None)
    tax = load_tax_records(fiscal_year if fiscal_year != "all" else None)
    return db, tax, fiscal_year, fys


@app.after_request
def no_cache(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return response


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        data = request.get_json() or {}
        username = data.get("username", "").strip().lower()
        password = data.get("password", "").strip()

        users = _load_users()
        user = users.get(username)
        if not user or user["password"] != password:
            return jsonify({"error": "Invalid username or password"}), 400

        session["user"] = {
            "username": user["username"],
            "name": user.get("name", username),
            "role": user.get("role", "User"),
            "department": user.get("department", "All Departments")
        }
        return jsonify({"success": True, "user": session["user"]})

    if "user" in session:
        return redirect(url_for("index"))
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login"))


@app.route("/api/me")
@login_required
def get_me():
    return jsonify(session.get("user"))


@app.route("/api/users", methods=["GET"])
@login_required
@admin_required
def list_users():
    users = _load_users()
    result = []
    for u, data in users.items():
        result.append({
            "username": data["username"],
            "name": data.get("name", ""),
            "role": data.get("role", "User"),
            "department": data.get("department", "All Departments")
        })
    return jsonify(result)


@app.route("/api/users", methods=["POST"])
@login_required
@admin_required
def create_user():
    data = request.get_json() or {}
    username = data.get("username", "").strip().lower()
    password = data.get("password", "").strip()
    name = data.get("name", "").strip()
    role = data.get("role", "Department Manager").strip()
    department = data.get("department", "BRAC IED").strip()

    if not username or not password:
        return jsonify({"error": "Username and Password are required"}), 400

    users = _load_users()
    if username in users:
        return jsonify({"error": f"User '{username}' already exists"}), 400

    users[username] = {
        "username": username,
        "password": password,
        "name": name or username.capitalize(),
        "role": role,
        "department": department
    }
    _save_users(users)
    log.info(f"Admin created user '{username}' with department scope '{department}'")
    return jsonify({"success": True, "user": users[username]})


@app.route("/api/users/<target_username>", methods=["PUT"])
@login_required
@admin_required
def update_user(target_username):
    data = request.get_json() or {}
    old_username = target_username.strip().lower()
    new_username = data.get("new_username", "").strip().lower() or old_username
    password = data.get("password", "").strip()
    name = data.get("name", "").strip()
    role = data.get("role", "").strip()
    department = data.get("department", "").strip()

    users = _load_users()
    if old_username not in users:
        return jsonify({"error": "Target user not found"}), 404

    if new_username != old_username and new_username in users:
        return jsonify({"error": f"Username '{new_username}' is already taken"}), 400

    user_info = users[old_username]

    if name:
        user_info["name"] = name
    if password:
        user_info["password"] = password
    if role:
        user_info["role"] = role
    if department:
        user_info["department"] = department

    if new_username != old_username:
        user_info["username"] = new_username
        del users[old_username]
        users[new_username] = user_info
        log.info(f"Admin renamed user '{old_username}' to '{new_username}'")
        current_user = session.get("user", {})
        if current_user.get("username") == old_username:
            session["user"] = user_info
    else:
        users[old_username] = user_info
        log.info(f"Admin updated user '{old_username}' details")
        current_user = session.get("user", {})
        if current_user.get("username") == old_username:
            session["user"] = user_info

    _save_users(users)
    return jsonify({"success": True, "message": f"Updated user '{new_username}'", "user": user_info})


@app.route("/api/user/profile", methods=["POST"])
@login_required
def update_my_profile():
    data = request.get_json() or {}
    current_user = session.get("user", {})
    username = current_user.get("username")
    if not username:
        return jsonify({"error": "Not authenticated"}), 401

    users = _load_users()
    if username not in users:
        return jsonify({"error": "User record not found"}), 404

    user_info = users[username]
    new_username = data.get("new_username", "").strip().lower() or username
    new_name = data.get("name", "").strip()
    new_password = data.get("new_password", "").strip()
    current_password = data.get("current_password", "").strip()

    if (new_password or new_username != username) and current_password != user_info.get("password"):
        return jsonify({"error": "Incorrect current password"}), 400

    if new_username != username and new_username in users:
        return jsonify({"error": f"Username '{new_username}' is already taken"}), 400

    if new_name:
        user_info["name"] = new_name
    if new_password:
        user_info["password"] = new_password

    if new_username != username:
        user_info["username"] = new_username
        del users[username]
        users[new_username] = user_info
    else:
        users[username] = user_info

    _save_users(users)
    session["user"] = user_info
    log.info(f"User '{username}' updated their profile")
    return jsonify({"success": True, "message": "Profile updated successfully", "user": user_info})


@app.route("/api/users/<username>", methods=["DELETE"])
@login_required
@admin_required
def delete_user(username):
    username = username.strip().lower()
    if username == "admin":
        return jsonify({"error": "Cannot delete primary admin user"}), 400

    users = _load_users()
    if username not in users:
        return jsonify({"error": "User not found"}), 404

    del users[username]
    _save_users(users)
    log.info(f"Admin deleted user '{username}'")
    return jsonify({"success": True})


@app.route("/")
@login_required
def index():
    return render_template("index.html")


def _get_scoped_employees(db, target_dept=None):
    user = session.get("user", {})
    username = user.get("username", "")
    user_dept = user.get("department", "All Departments")
    user_role = user.get("role", "")

    has_full_access = (user_dept == "All Departments") or ("Full Access" in user_role) or (user_role == "Super Admin") or (username in ["admin", "accounts", "hr"])

    dept_filter = target_dept or request.args.get("department")

    if has_full_access and not dept_filter:
        return db

    filtered = {}
    for pin, emp in db.items():
        if not has_full_access and emp.department != user_dept and user_dept not in emp.department:
            continue
        if dept_filter and dept_filter != "All Departments" and emp.department != dept_filter and dept_filter not in emp.department:
            continue
        filtered[pin] = emp

    return filtered if filtered else db


@app.route("/api/fiscal-years")
@login_required
def list_fiscal_years():
    fy_param = request.args.get("fiscal_year")
    db, tax, active_fy, fys = _get_stores(fy_param)
    return jsonify({
        "fiscal_years": fys,
        "active_fiscal_year": active_fy
    })


@app.route("/api/status")
@login_required
def api_status():
    meta = _load_meta()
    fy_param = request.args.get("fiscal_year")
    db, tax, active_fy, fys = _get_stores(fy_param)
    scoped_db = _get_scoped_employees(db)
    depts = sorted(list(set(e.department for e in db.values() if e.department)))
    if "All Departments" not in depts:
        depts.insert(0, "All Departments")

    return jsonify({
        "active_fiscal_year": active_fy,
        "fiscal_years": fys,
        "database_count": len(scoped_db),
        "total_database_count": len(db),
        "tax_count": sum(len(tax.get(pin, [])) for pin in scoped_db),
        "tax_pins": sum(1 for pin in scoped_db if pin in tax),
        "departments": depts,
        "user": session.get("user"),
        "meta": meta,
    })


@app.route("/api/stats")
@login_required
def api_stats():
    fy_param = request.args.get("fiscal_year")
    db, tax, active_fy, fys = _get_stores(fy_param)
    scoped_db = _get_scoped_employees(db)
    total_emp = len(scoped_db)
    male_count = sum(1 for e in scoped_db.values() if e.gender.lower() == 'male')
    female_count = sum(1 for e in scoped_db.values() if e.gender.lower() == 'female')
    total_gross = sum(e.gross for e in scoped_db.values())
    total_net = sum(e.net_total for e in scoped_db.values())

    total_tax_claimed = 0.0
    tax_emp_count = 0
    total_challan_count = 0

    for pin in scoped_db:
        recs = tax.get(pin, [])
        if recs:
            tax_emp_count += 1
            total_challan_count += len(recs)
            total_tax_claimed += sum(r.claim_amount for r in recs)

    return jsonify({
        "active_fiscal_year": active_fy,
        "fiscal_years": fys,
        "total_employees": total_emp,
        "male_count": male_count,
        "female_count": female_count,
        "total_gross": total_gross,
        "total_net": total_net,
        "total_tax_claimed": total_tax_claimed,
        "tax_emp_count": tax_emp_count,
        "total_challans": total_challan_count,
        "tax_compliance_pct": round((tax_emp_count / total_emp * 100), 1) if total_emp > 0 else 0
    })


@app.route("/api/audit")
@login_required
def api_audit():
    fy_param = request.args.get("fiscal_year")
    db, tax, active_fy, fys = _get_stores(fy_param)
    scoped_db = _get_scoped_employees(db)
    issues = []

    db_without_tax_count = 0
    unmatched_tax_pins_count = 0
    salary_anomaly_count = 0

    for pin, emp in scoped_db.items():
        if emp.net_total <= 0:
            salary_anomaly_count += 1
            issues.append({
                "type": "salary_anomaly",
                "severity": "warning",
                "pin": pin,
                "name": emp.name or "N/A",
                "department": emp.department or "N/A",
                "message": f"Zero or negative net salary (BDT {fmt_bdt(emp.net_total)})."
            })

        if tax and pin not in tax:
            db_without_tax_count += 1
            issues.append({
                "type": "db_without_tax",
                "severity": "info",
                "pin": pin,
                "name": emp.name or "N/A",
                "department": emp.department or "N/A",
                "message": "No tax challan records found for this employee."
            })

    if tax:
        for pin, recs in tax.items():
            if pin not in db:
                unmatched_tax_pins_count += 1
                tot_claim = sum(r.claim_amount for r in recs)
                issues.append({
                    "type": "tax_without_db",
                    "severity": "danger",
                    "pin": pin,
                    "name": "Unmatched PIN",
                    "department": "N/A",
                    "message": f"{len(recs)} Tax Challan record(s) totaling BDT {fmt_bdt(tot_claim)} found for PIN {pin}, but PIN is missing from Employee Database."
                })

    return jsonify({
        "summary": {
            "active_fiscal_year": active_fy,
            "total_db_employees": len(scoped_db),
            "total_tax_pins": len(tax),
            "total_issues": len(issues),
            "unmatched_tax_pins_count": unmatched_tax_pins_count,
            "db_without_tax_count": db_without_tax_count,
            "salary_anomaly_count": salary_anomaly_count,
        },
        "issues": issues
    })


_cached_employee_list = {}


def _decode_file_content(file_bytes: bytes) -> str:
    for enc in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
        try:
            return file_bytes.decode(enc)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="replace")


@app.route("/api/datasets")
@login_required
def list_datasets():
    datasets = get_all_dataset_batches()
    return jsonify({"datasets": datasets})


@app.route("/api/dataset/<batch_id>", methods=["DELETE"])
@login_required
def delete_dataset_batch(batch_id):
    success = delete_batch(batch_id)
    if success:
        return jsonify({"success": True, "message": f"Dataset batch '{batch_id}' removed successfully"})
    return jsonify({"error": "Batch not found"}), 404


@app.route("/api/upload/database", methods=["POST"])
@login_required
def upload_database():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    filename = f.filename
    clean_filename = f.filename.lower()
    if not (clean_filename.endswith(".csv") or clean_filename.endswith(".xlsx") or clean_filename.endswith(".xls")):
        return jsonify({"error": "Only CSV or Excel (.xlsx/.xls) files supported"}), 400

    try:
        user = session.get("user", {})
        username = user.get("username", "admin")

        if clean_filename.endswith(".xlsx") or clean_filename.endswith(".xls"):
            file_bytes = f.read()
            with open(DB_EXCEL_PATH, "wb") as df:
                df.write(file_bytes)
            result = parse_database_excel(file_bytes)
        else:
            file_bytes = f.read()
            content = _decode_file_content(file_bytes)
            with open(DB_CSV_PATH, "w", encoding="utf-8") as df:
                df.write(content)
            result = parse_database_csv(content)

        sample_emp = next(iter(result.values()), None)
        fy = f"{sample_emp.start_year}-{sample_emp.end_year}" if sample_emp else "2025-2026"

        save_employees(result, default_fiscal_year=fy, filename=filename, uploaded_by=username)
        _save_meta("database", filename, len(result))
        log.info(f"Database uploaded: {len(result)} employees for FY {fy} by {username} from {filename}")
        return jsonify({"count": len(result), "fiscal_year": fy, "filename": filename, "pins": list(result.keys())})
    except Exception as e:
        log.error(f"DB parse error: {traceback.format_exc()}")
        return jsonify({"error": str(e)}), 400


@app.route("/api/upload/tax", methods=["POST"])
@login_required
def upload_tax():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    filename = f.filename
    clean_filename = f.filename.lower()
    if not (clean_filename.endswith(".csv") or clean_filename.endswith(".xlsx") or clean_filename.endswith(".xls")):
        return jsonify({"error": "Only CSV or Excel (.xlsx/.xls) files supported"}), 400

    try:
        user = session.get("user", {})
        username = user.get("username", "admin")

        fy_param = request.args.get("fiscal_year")
        if not fy_param or fy_param == "all":
            fys = get_fiscal_years()
            fy_param = fys[0] if fys else "2025-2026"

        if clean_filename.endswith(".xlsx") or clean_filename.endswith(".xls"):
            file_bytes = f.read()
            with open(TAX_EXCEL_PATH, "wb") as df:
                df.write(file_bytes)
            result = parse_tax_excel(file_bytes)
        else:
            file_bytes = f.read()
            content = _decode_file_content(file_bytes)
            with open(TAX_CSV_PATH, "w", encoding="utf-8") as df:
                df.write(content)
            result = parse_tax_csv(content)

        save_tax_records(result, default_fiscal_year=fy_param, filename=filename, uploaded_by=username)
        total = sum(len(v) for v in result.values())
        _save_meta("tax", filename, total)
        log.info(f"Tax uploaded: {total} records for FY {fy_param} by {username} from {filename}")
        return jsonify({"count": total, "fiscal_year": fy_param, "filename": filename, "pins": list(result.keys())})
    except Exception as e:
        log.error(f"Tax parse error: {traceback.format_exc()}")
        return jsonify({"error": str(e)}), 400


def _clear_meta(key):
    meta = _load_meta()
    if key in meta:
        del meta[key]
        with open(META_PATH, "w") as f:
            json.dump(meta, f, indent=2)


@app.route("/api/upload/<file_type>", methods=["DELETE"])
@login_required
def remove_uploaded_file(file_type):
    fy = request.args.get("fiscal_year")
    if file_type == "database":
        delete_fiscal_year(fy) if (fy and fy.lower() != 'all') else clear_all_data()
        _clear_meta("database")
        for p in [DB_CSV_PATH, DB_EXCEL_PATH]:
            if os.path.exists(p):
                try:
                    os.remove(p)
                except Exception:
                    pass
        return jsonify({"success": True, "message": f"Database records for FY {fy or 'All'} removed"})
    elif file_type == "tax":
        if fy and fy.lower() != 'all':
            with get_connection() as conn:
                conn.cursor().execute("DELETE FROM tax_records WHERE fiscal_year = ?", (fy,))
                conn.commit()
        else:
            with get_connection() as conn:
                conn.cursor().execute("DELETE FROM tax_records")
                conn.commit()
        _clear_meta("tax")
        for p in [TAX_CSV_PATH, TAX_EXCEL_PATH]:
            if os.path.exists(p):
                try:
                    os.remove(p)
                except Exception:
                    pass
        return jsonify({"success": True, "message": f"Tax records for FY {fy or 'All'} removed"})
    elif file_type == "all":
        clear_all_data()
        _clear_meta("all")
        for p in [DB_CSV_PATH, DB_EXCEL_PATH, TAX_CSV_PATH, TAX_EXCEL_PATH]:
            if os.path.exists(p):
                try:
                    os.remove(p)
                except Exception:
                    pass
        return jsonify({"success": True, "message": "All database and tax records removed"})
    return jsonify({"error": "Invalid file type"}), 400


@app.route("/api/employees")
@login_required
def list_employees():
    fy_param = request.args.get("fiscal_year")
    db, tax, active_fy, fys = _get_stores(fy_param)
    scoped_db = _get_scoped_employees(db)
    employees = []
    for pin, emp in scoped_db.items():
        tax_recs = tax.get(pin, [])
        total_claim = sum(r.claim_amount for r in tax_recs)
        total_challan = sum(r.total_challan_amount for r in tax_recs)
        employees.append({
            **emp.to_dict(),
            "fiscal_year": active_fy,
            "tax_records_count": len(tax_recs),
            "total_claim": total_claim,
            "total_challan": total_challan,
        })
    return jsonify(employees)


@app.route("/api/employee/<path:pin>")
def get_employee(pin):
    pin = pin.strip()
    fy_param = request.args.get("fiscal_year")
    db, tax, active_fy, fys = _get_stores(fy_param)
    emp = db.get(pin)
    if not emp:
        for k in db:
            if k.strip() == pin:
                emp = db[k]
                pin = k
                break
    if not emp:
        return jsonify({"error": "Employee not found"}), 404
    tax_recs = tax.get(pin, [])
    if not tax_recs:
        for k in tax:
            if k.strip() == pin:
                tax_recs = tax[k]
                break
    return jsonify({
        "employee": emp.to_dict(),
        "tax_records": [r.to_dict() for r in tax_recs],
    })


@app.route("/api/employee/<path:pin>/update", methods=["POST"])
@login_required
def update_employee_route(pin):
    pin = pin.strip()
    data = request.get_json() or {}
    tin = data.get("tin")
    name = data.get("name")
    designation = data.get("designation")
    fy = request.args.get("fiscal_year") or data.get("fiscal_year")

    success = update_employee_tin(pin, fiscal_year=fy, new_tin=tin, new_name=name, new_designation=designation)
    if success:
        log.info(f"Updated employee PIN {pin}: tin='{tin}', name='{name}', designation='{designation}' for FY {fy or 'All'}")
        return jsonify({"success": True, "message": f"Updated employee info for PIN {pin}"})
    return jsonify({"error": f"Employee PIN {pin} not found or no changes made"}), 400


@app.route("/api/generate/<path:pin>")
def generate_single(pin):
    pin = pin.strip()
    fy_param = request.args.get("fiscal_year")
    db, tax, active_fy, fys = _get_stores(fy_param)
    emp = db.get(pin)
    if not emp:
        for k in db:
            if k.strip() == pin:
                emp = db[k]
                pin = k
                break
    if not emp:
        return jsonify({"error": "Employee not found"}), 404
    name = request.args.get("name", emp.name or "")
    designation = request.args.get("designation", emp.designation or "")
    tin_arg = request.args.get("tin")
    if tin_arg and tin_arg.strip():
        emp.tin = tin_arg.strip()
        update_employee_tin(pin, fiscal_year=active_fy, new_tin=emp.tin)

    use_letterhead = request.args.get("letterhead") in ["1", "true", "True"]
    start_year = request.args.get("start_year", type=int) or getattr(emp, "start_year", 2025)
    end_year = request.args.get("end_year", type=int) or getattr(emp, "end_year", 2026)
    tax_recs = tax.get(pin, [])
    batch_id = "single"
    batch_dir = os.path.join(OUTPUT_DIR, batch_id)
    os.makedirs(batch_dir, exist_ok=True)
    filename = f"IncomeTax_{pin}.pdf"
    filepath = os.path.join(batch_dir, filename)
    generate_pdf(emp, tax_recs, filepath, name=name, designation=designation, start_year=start_year, end_year=end_year, use_letterhead=use_letterhead)
    return jsonify({"batch_id": batch_id, "filename": filename, "pin": pin})


@app.route("/api/preview/<path:pin>")
def preview_single(pin):
    pin = pin.strip()
    fy_param = request.args.get("fiscal_year")
    db, tax, active_fy, fys = _get_stores(fy_param)
    emp = db.get(pin)
    if not emp:
        for k in db:
            if k.strip() == pin:
                emp = db[k]
                pin = k
                break
    if not emp:
        return jsonify({"error": "Employee not found"}), 404
    name = request.args.get("name", emp.name or "")
    designation = request.args.get("designation", emp.designation or "")
    tin_arg = request.args.get("tin")
    if tin_arg and tin_arg.strip():
        emp.tin = tin_arg.strip()
        update_employee_tin(pin, fiscal_year=active_fy, new_tin=emp.tin)

    use_letterhead = request.args.get("letterhead") in ["1", "true", "True"]
    start_year = request.args.get("start_year", type=int) or getattr(emp, "start_year", 2025)
    end_year = request.args.get("end_year", type=int) or getattr(emp, "end_year", 2026)
    tax_recs = tax.get(pin, [])

    preview_dir = os.path.join(OUTPUT_DIR, "preview")
    os.makedirs(preview_dir, exist_ok=True)
    filepath = os.path.join(preview_dir, f"preview_{pin}.pdf")
    generate_pdf(emp, tax_recs, filepath, name=name, designation=designation, start_year=start_year, end_year=end_year, use_letterhead=use_letterhead)
    return send_file(filepath, mimetype="application/pdf", as_attachment=False)


@app.route("/api/settings/letterhead", methods=["POST"])
@login_required
@admin_required
def upload_letterhead():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    if not file or not file.filename:
        return jsonify({"error": "Empty file"}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in [".png", ".jpg", ".jpeg"]:
        return jsonify({"error": "Only PNG or JPG images are allowed for letterhead"}), 400

    save_path = os.path.join(DATA_DIR, "custom_letterhead.png")
    file.save(save_path)
    log.info("Custom letterhead image uploaded successfully")
    return jsonify({"success": True, "message": "Custom letterhead uploaded successfully"})


@app.route("/api/generate-all")
def generate_all():
    fy_param = request.args.get("fiscal_year")
    db, tax, active_fy, fys = _get_stores(fy_param)
    scoped_db = _get_scoped_employees(db)
    if not scoped_db:
        return jsonify({"error": "No database loaded for active fiscal year"}), 400

    batch_id = str(uuid.uuid4())[:8]
    batch_dir = os.path.join(OUTPUT_DIR, batch_id)
    os.makedirs(batch_dir, exist_ok=True)
    generated, errors = [], []
    for pin, emp in scoped_db.items():
        try:
            tax_recs = tax.get(pin, [])
            filename = f"IncomeTax_{pin}.pdf"
            filepath = os.path.join(batch_dir, filename)
            start_year = getattr(emp, "start_year", 2025)
            end_year = getattr(emp, "end_year", 2026)
            generate_pdf(emp, tax_recs, filepath, name=emp.name, designation=emp.designation, start_year=start_year, end_year=end_year)
            generated.append({"pin": pin, "filename": filename})
        except Exception as e:
            errors.append({"pin": pin, "error": str(e)})
    return jsonify({"batch_id": batch_id, "total": len(scoped_db),
                    "generated": generated, "errors": errors})


@app.route("/api/export/excel")
def export_excel():
    fy_param = request.args.get("fiscal_year")
    db, tax, active_fy, fys = _get_stores(fy_param)
    scoped_db = _get_scoped_employees(db)
    if not scoped_db:
        return jsonify({"error": "No database loaded to export for active fiscal year"}), 400

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    wb = openpyxl.Workbook()

    # Sheet 1: Employee Summary
    ws_emp = wb.active
    ws_emp.title = "Employee Salary & Tax Summary"

    headers_emp = [
        "PIN", "Name", "Designation", "Gender", "TIN",
        "Basic (50%)", "House Rent (30%)", "Medical (10%)", "Conveyance (10%)",
        "Festival Bonus", "Arrears", "Others", "Gross Salary", "Net Total",
        "Tax Challans Count", "Total Tax Deducted/Claimed"
    ]

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    ws_emp.append(headers_emp)
    for col_num, h in enumerate(headers_emp, 1):
        cell = ws_emp.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    row_idx = 2
    for pin, emp in scoped_db.items():
        recs = tax.get(pin, [])
        tax_claimed = sum(r.claim_amount for r in recs)

        row_data = [
            emp.pin, emp.name, emp.designation, emp.gender, emp.tin,
            emp.basic, emp.house_rent, emp.medical_allowance, emp.conveyance_allowance,
            emp.festival_bonus, emp.arrears, emp.others, emp.gross, emp.net_total,
            len(recs), tax_claimed
        ]
        ws_emp.append(row_data)

        ws_emp.cell(row=row_idx, column=1).alignment = center_align
        ws_emp.cell(row=row_idx, column=2).alignment = left_align
        ws_emp.cell(row=row_idx, column=3).alignment = left_align
        ws_emp.cell(row=row_idx, column=4).alignment = center_align
        ws_emp.cell(row=row_idx, column=5).alignment = center_align

        for c in range(6, 15):
            ws_emp.cell(row=row_idx, column=c).number_format = '#,##0'
            ws_emp.cell(row=row_idx, column=c).alignment = right_align

        ws_emp.cell(row=row_idx, column=15).alignment = center_align
        ws_emp.cell(row=row_idx, column=16).number_format = '#,##0'
        ws_emp.cell(row=row_idx, column=16).alignment = right_align

        for col_num in range(1, len(headers_emp) + 1):
            ws_emp.cell(row=row_idx, column=col_num).border = thin_border

        row_idx += 1

    for col in ws_emp.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_emp.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Sheet 2: Tax Challans Detail
    ws_tax = wb.create_sheet(title="Tax Challan Records")
    headers_tax = ["PIN", "Month", "A-Challan No", "Challan Date", "Claim Amount", "Total Challan Amount", "Bank Information"]

    ws_tax.append(headers_tax)
    for col_num, h in enumerate(headers_tax, 1):
        cell = ws_tax.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.alignment = center_align

    tax_row_idx = 2
    for pin, recs in tax.items():
        if pin not in scoped_db and active_fy and active_fy.lower() != 'all':
            continue
        for r in recs:
            ws_tax.append([
                r.pin, r.month, r.challan_no, r.challan_date,
                r.claim_amount, r.total_challan_amount, r.bank_info
            ])
            ws_tax.cell(row=tax_row_idx, column=1).alignment = center_align
            ws_tax.cell(row=tax_row_idx, column=2).alignment = center_align
            ws_tax.cell(row=tax_row_idx, column=3).alignment = center_align
            ws_tax.cell(row=tax_row_idx, column=4).alignment = center_align
            ws_tax.cell(row=tax_row_idx, column=5).number_format = '#,##0'
            ws_tax.cell(row=tax_row_idx, column=5).alignment = right_align
            ws_tax.cell(row=tax_row_idx, column=6).number_format = '#,##0'
            ws_tax.cell(row=tax_row_idx, column=6).alignment = right_align
            ws_tax.cell(row=tax_row_idx, column=7).alignment = left_align

            for col_num in range(1, len(headers_tax) + 1):
                ws_tax.cell(row=tax_row_idx, column=col_num).border = thin_border
            tax_row_idx += 1

    for col in ws_tax.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_tax.column_dimensions[col_letter].width = max(max_len + 4, 14)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"IncomeTax_Summary_{timestamp}.xlsx"
    )


@app.route("/download/<batch_id>/<filename>")
def download(batch_id, filename):
    d = os.path.join(OUTPUT_DIR, secure_filename(batch_id))
    return send_from_directory(d, secure_filename(filename), as_attachment=True)


@app.route("/download-all/<batch_id>")
def download_all(batch_id):
    import zipfile, io
    d = os.path.join(OUTPUT_DIR, secure_filename(batch_id))
    if not os.path.isdir(d):
        return jsonify({"error": "Not found"}), 404
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname in os.listdir(d):
            if fname.endswith(".pdf"):
                zf.write(os.path.join(d, fname), fname)
    buf.seek(0)
    return send_file(buf, mimetype="application/zip",
                     as_attachment=True, download_name=f"IncomeTax_{batch_id}.zip")


SMTP_CONFIG_PATH = os.path.join(DATA_DIR, "smtp.json")

def _load_smtp_config():
    if os.path.exists(SMTP_CONFIG_PATH):
        try:
            with open(SMTP_CONFIG_PATH) as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _send_email(to_email: str, subject: str, body_text: str) -> bool:
    config = _load_smtp_config()
    smtp_server = config.get("smtp_server")
    smtp_port = int(config.get("smtp_port", 587))
    smtp_user = config.get("smtp_user")
    smtp_password = config.get("smtp_password")
    sender_email = config.get("sender_email", smtp_user or "noreply@bracied.edu.bd")

    if not smtp_server or not smtp_user:
        log.info(f"[SIMULATED EMAIL NOTIFICATION] To: {to_email} | Subject: {subject}\n{body_text}")
        return True

    try:
        msg = MIMEMultipart()
        msg["From"] = sender_email
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body_text, "plain"))

        server = smtplib.SMTP(smtp_server, smtp_port, timeout=10)
        server.starttls()
        if smtp_password:
            server.login(smtp_user, smtp_password)
        server.sendmail(sender_email, [to_email], msg.as_string())
        server.quit()
        log.info(f"Email sent successfully to {to_email}")
        return True
    except Exception as e:
        log.error(f"Failed to send email to {to_email}: {e}")
        return False


@app.route("/api/request-statement", methods=["POST"])
def submit_tax_request():
    data = request.get_json() or {}
    pin = data.get("pin", "").strip()
    name = data.get("name", "").strip()
    designation = data.get("designation", "").strip()
    tin = data.get("tin", "").strip()
    email = data.get("email", "").strip().lower()
    fiscal_year = data.get("fiscal_year", "").strip() or "2025-2026"
    remarks = data.get("remarks", "").strip()

    if not pin or not name or not email:
        return jsonify({"error": "Employee PIN, Full Name, and Email Address are required"}), 400

    emp = None
    try:
        db, _, _, _ = _get_stores(fiscal_year)
        emp = db.get(pin)
    except Exception:
        pass

    if emp:
        if not designation and getattr(emp, 'designation', None):
            designation = emp.designation
        if not tin and getattr(emp, 'tin', None):
            tin = emp.tin

    req = create_tax_request(pin, name, designation, tin, email, fiscal_year, remarks)
    
    subject = f"[BRAC IED] Income Tax Certificate Request Received ({req['req_id']})"
    body = (
        f"Dear {name},\n\n"
        f"Your request for Income Tax Statement for Financial Year {fiscal_year} has been received by HR.\n\n"
        f"Request Details:\n"
        f"- Request ID: {req['req_id']}\n"
        f"- PIN Number: {pin}\n"
        f"- Designation: {designation or 'N/A'}\n"
        f"- Status: Pending HR Review\n\n"
        f"You will receive another email notification when your statement is prepared, printed, and ready for pickup at HR.\n\n"
        f"Regards,\nBRAC Institute of Educational Development (BRAC IED)"
    )
    _send_email(email, subject, body)

    return jsonify({"success": True, "request": req, "message": f"Request submitted successfully! Tracking ID: {req['req_id']}"})


@app.route("/api/request-statement/track/<path:query>")
def track_tax_request(query):
    req = get_tax_request_by_req_id(query)
    if not req:
        return jsonify({"error": "Request not found for specified ID or PIN"}), 404
    return jsonify({"request": req})


@app.route("/api/requests")
@login_required
def list_tax_requests():
    fy = request.args.get("fiscal_year")
    st = request.args.get("status")
    requests_list = get_tax_requests(fiscal_year=fy, status=st)

    all_reqs = get_tax_requests(fiscal_year=fy)
    summary = {
        "total": len(all_reqs),
        "pending": sum(1 for r in all_reqs if r["status"] == "Pending"),
        "in_progress": sum(1 for r in all_reqs if r["status"] in ("In Progress", "Preparing")),
        "printed": sum(1 for r in all_reqs if r["status"] in ("Printed", "Statement Prepared")),
        "ready": sum(1 for r in all_reqs if r["status"] == "Ready for Pickup"),
        "completed": sum(1 for r in all_reqs if r["status"] == "Completed"),
    }
    return jsonify({"requests": requests_list, "summary": summary})


@app.route("/api/request/<req_id>/status", methods=["PUT"])
@login_required
def update_request_status_api(req_id):
    data = request.get_json() or {}
    new_status = data.get("status", "").strip()
    hr_notes = data.get("hr_notes", "").strip()
    notify_email = data.get("notify_email", True)

    user = session.get("user", {})
    prepared_by = user.get("name") or user.get("username") or "HR Admin"

    if not new_status:
        return jsonify({"error": "Status is required"}), 400

    req = update_tax_request_status(req_id, new_status, hr_notes=hr_notes, prepared_by=prepared_by)
    if not req:
        return jsonify({"error": "Request not found"}), 404

    if notify_email or new_status == "Ready for Pickup":
        if new_status == "Ready for Pickup":
            subject = f"[BRAC IED] Your Income Tax Statement for FY {req['fiscal_year']} is READY FOR PICKUP!"
            body = (
                f"Dear {req['name']},\n\n"
                f"Great news! Your Income Tax Certificate Statement for Financial Year {req['fiscal_year']} (PIN: {req['pin']}) "
                f"has been prepared, printed, and verified by HR.\n\n"
                f"Pickup Location: HR Department, BRAC Institute of Educational Development\n"
                f"Tracking ID: {req['req_id']}\n"
                f"Status: Ready for Pickup\n\n"
                f"Please visit the HR office to collect your physical signed copy.\n\n"
                f"Regards,\nHR & Payroll Team\nBRAC Institute of Educational Development (BRAC IED)"
            )
        elif new_status in ("In Progress", "Preparing"):
            subject = f"[BRAC IED] HR is Preparing Your Income Tax Statement ({req['req_id']})"
            body = (
                f"Dear {req['name']},\n\n"
                f"HR is currently preparing your Income Tax Certificate Statement for FY {req['fiscal_year']} (PIN: {req['pin']}).\n"
                f"Status: In Progress / Preparing\n\n"
                f"You will receive an update as soon as the physical statement is printed and ready.\n\n"
                f"Regards,\nBRAC IED HR Team"
            )
        elif new_status == "Completed":
            subject = f"[BRAC IED] Income Tax Statement Request Completed ({req['req_id']})"
            body = (
                f"Dear {req['name']},\n\n"
                f"Your request for Income Tax Statement for FY {req['fiscal_year']} (PIN: {req['pin']}) has been marked as COMPLETED.\n"
                f"Thank you!\n\n"
                f"Regards,\nBRAC IED HR Team"
            )
        else:
            subject = f"[BRAC IED] Status Update on Tax Statement Request ({req['req_id']})"
            body = (
                f"Dear {req['name']},\n\n"
                f"The status of your Income Tax Statement Request ({req['req_id']}) has been updated to: {new_status}.\n"
                f"Notes: {hr_notes or 'None'}\n\n"
                f"Regards,\nBRAC IED HR Team"
            )
        _send_email(req['email'], subject, body)

    return jsonify({"success": True, "request": req, "message": f"Status updated to '{new_status}'"})


@app.route("/api/settings/smtp", methods=["GET", "POST"])
@login_required
def manage_smtp_settings():
    if request.method == "POST":
        data = request.get_json() or {}
        with open(SMTP_CONFIG_PATH, "w") as f:
            json.dump(data, f, indent=2)
        return jsonify({"success": True, "message": "SMTP settings saved successfully"})
    else:
        config = _load_smtp_config()
        config["smtp_password"] = "******" if config.get("smtp_password") else ""
        return jsonify({"config": config})


_seed_db_if_empty()

if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=80)

